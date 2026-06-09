import csv
import re
import unicodedata
from datetime import timedelta
from io import BytesIO

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseForbidden, HttpResponseBadRequest
from django.utils import timezone
from django.core.paginator import Paginator
from django.db.models import Count
from django.db.models.functions import TruncDate

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from .models import Informe
from core_apps.camera.models import SecurityEvent
from core_apps.common.permissions import is_admin_user


RISK_LEVELS = ("BAJO", "MEDIO", "ALTO", "CRITICO")
DEFAULT_EVENT_LEVELS = {
    "face_recognized": "BAJO",
    "face_unknown": "MEDIO",
    "ppe_missing": "ALTO",
    "intrusion": "ALTO",
    "authorized_object": "BAJO",
    "unauthorized_object": "MEDIO",
    "dangerous_object": "ALTO",
    "unauthorized_access": "ALTO",
}
RISK_LEVEL_PATTERN = re.compile(
    r"(?:nivel|prioridad)\s*:?\s*(BAJO|BAJA|MEDIO|MEDIA|ALTO|ALTA|CRITICO)",
    re.IGNORECASE,
)


def _strip_accents(value):
    return "".join(
        char
        for char in unicodedata.normalize("NFKD", str(value))
        if not unicodedata.combining(char)
    )


def _normalize_risk_level(value, default="MEDIO"):
    if not value:
        return default

    level = _strip_accents(value).strip().upper()

    aliases = {
        "BAJA": "BAJO",
        "MEDIA": "MEDIO",
        "ALTA": "ALTO",
    }
    level = aliases.get(level, level)

    if level in RISK_LEVELS:
        return level

    return default


def _extract_risk_level(details, event_type=None):
    if hasattr(details, "severity") and details.severity:
        return _normalize_risk_level(details.severity)

    normalized_details = _strip_accents(details or "")
    match = RISK_LEVEL_PATTERN.search(normalized_details)

    if match:
        return _normalize_risk_level(match.group(1))

    return DEFAULT_EVENT_LEVELS.get(event_type, "MEDIO")


def _add_relative_percentages(items):
    max_total = max((item["total"] for item in items), default=0)

    for item in items:
        item["percent"] = round((item["total"] / max_total) * 100) if max_total else 0

    return items


def _get_event_type_counts(events_qs):
    event_type_display = dict(SecurityEvent.EVENT_TYPES)
    items = [
        {
            "key": item["event_type"],
            "label": event_type_display.get(item["event_type"], item["event_type"]),
            "total": item["total"],
        }
        for item in (
            events_qs
            .values("event_type")
            .annotate(total=Count("id"))
            .order_by("-total", "event_type")
        )
    ]

    return _add_relative_percentages(items)


def _get_event_date_counts(events_qs):
    items = [
        {
            "label": item["day"].strftime("%d/%m/%Y") if item["day"] else "Sin fecha",
            "total": item["total"],
        }
        for item in (
            events_qs
            .annotate(day=TruncDate("timestamp", tzinfo=timezone.get_current_timezone()))
            .values("day")
            .annotate(total=Count("id"))
            .order_by("-day")
        )
    ]

    return _add_relative_percentages(items)


def _get_risk_level_counts(events_qs):
    counters = {level: 0 for level in RISK_LEVELS}

    for event in events_qs.values("event_type", "details", "severity").iterator():
        level = (
            _normalize_risk_level(event["severity"])
            if event.get("severity")
            else _extract_risk_level(event["details"], event["event_type"])
        )
        counters[level] = counters.get(level, 0) + 1

    total_events = sum(counters.values())
    labels = {
        "BAJO": "Bajo",
        "MEDIO": "Medio",
        "ALTO": "Alto",
        "CRITICO": "Crítico",
    }

    return [
        {
            "key": level.lower(),
            "label": labels[level],
            "total": counters[level],
            "percent": round((counters[level] / total_events) * 100) if total_events else 0,
        }
        for level in RISK_LEVELS
    ]


@login_required(login_url="/login/")
def lista_informes(request):
    if not is_admin_user(request.user):
        return HttpResponseForbidden("Solo un administrador puede ver informes.")

    informes_qs = Informe.objects.select_related('security_event').order_by('-fecha')
    events_qs = SecurityEvent.objects.all()

    total_count = informes_qs.count()
    critical_count = informes_qs.filter(epp_correcto=False).count()
    event_type_counts = _get_event_type_counts(events_qs)
    event_date_counts = _get_event_date_counts(events_qs)
    risk_level_counts = _get_risk_level_counts(events_qs)

    paginator = Paginator(informes_qs, 9)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)
    page_range = paginator.get_elided_page_range(
        number=page_obj.number,
        on_each_side=1,
        on_ends=1,
    )

    return render(request, 'informes/index.html', {
        'informes': page_obj,
        'page_obj': page_obj,
        'page_range': page_range,
        'total_count': total_count,
        'critical_count': critical_count,
        'event_type_counts': event_type_counts,
        'event_date_counts': event_date_counts,
        'risk_level_counts': risk_level_counts,
        'segment': 'lista_informes',
    })


def obtener_rango_fechas(periodo):
    """
    Retorna el rango de fechas según el período seleccionado.
    semanal: semana actual
    mensual: mes actual
    anual: año actual
    """
    ahora = timezone.localtime()
    inicio_hoy = ahora.replace(hour=0, minute=0, second=0, microsecond=0)

    if periodo == "semanal":
        inicio = inicio_hoy - timedelta(days=inicio_hoy.weekday())
        nombre_periodo = "Semanal"

    elif periodo == "mensual":
        inicio = inicio_hoy.replace(day=1)
        nombre_periodo = "Mensual"

    elif periodo == "anual":
        inicio = inicio_hoy.replace(month=1, day=1)
        nombre_periodo = "Anual"

    else:
        return None, None, None

    fin = ahora
    return inicio, fin, nombre_periodo


def obtener_informes_por_periodo(periodo):
    inicio, fin, nombre_periodo = obtener_rango_fechas(periodo)

    if not inicio:
        return None, None, None, None

    informes = Informe.objects.filter(
        fecha__gte=inicio,
        fecha__lte=fin
    ).order_by('-fecha')

    return informes, inicio, fin, nombre_periodo


def estado_epp(informe):
    return "EPP correcto" if informe.epp_correcto else "Alerta / Sin EPP"


@login_required(login_url="/login/")
def exportar_informes(request, periodo, formato):
    if not is_admin_user(request.user):
        return HttpResponseForbidden("Solo un administrador puede generar reportes.")

    formatos_permitidos = ["csv", "xlsx", "pdf"]

    if formato not in formatos_permitidos:
        return HttpResponseBadRequest("Formato de reporte no válido.")

    informes, inicio, fin, nombre_periodo = obtener_informes_por_periodo(periodo)

    if informes is None:
        return HttpResponseBadRequest("Período de reporte no válido.")

    fecha_actual = timezone.localtime().strftime("%Y%m%d_%H%M%S")
    nombre_archivo = f"reporte_informes_{periodo}_{fecha_actual}"

    if formato == "csv":
        return generar_reporte_csv(informes, nombre_archivo, nombre_periodo, inicio, fin)

    if formato == "xlsx":
        return generar_reporte_excel(informes, nombre_archivo, nombre_periodo, inicio, fin)

    if formato == "pdf":
        return generar_reporte_pdf(informes, nombre_archivo, nombre_periodo, inicio, fin)


def generar_reporte_csv(informes, nombre_archivo, nombre_periodo, inicio, fin):
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}.csv"'

    # BOM para que Excel abra correctamente caracteres con tilde y ñ
    response.write("\ufeff")

    writer = csv.writer(response)
    writer.writerow([f"Reporte {nombre_periodo} de Informes"])
    writer.writerow([
        "Desde",
        timezone.localtime(inicio).strftime("%d/%m/%Y %H:%M:%S"),
        "Hasta",
        timezone.localtime(fin).strftime("%d/%m/%Y %H:%M:%S")
    ])
    writer.writerow([])
    writer.writerow(["Fecha", "Cámara", "Persona detectada", "Estado EPP", "Descripción"])

    for informe in informes:
        writer.writerow([
            timezone.localtime(informe.fecha).strftime("%d/%m/%Y %H:%M:%S"),
            informe.camara,
            informe.persona_detectada or "No identificada",
            estado_epp(informe),
            informe.descripcion or ""
        ])

    return response


def generar_reporte_excel(informes, nombre_archivo, nombre_periodo, inicio, fin):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Reporte {nombre_periodo}"

    ws.merge_cells("A1:E1")
    ws["A1"] = f"Reporte {nombre_periodo} de Informes"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = "Desde:"
    ws["B2"] = timezone.localtime(inicio).strftime("%d/%m/%Y %H:%M:%S")
    ws["D2"] = "Hasta:"
    ws["E2"] = timezone.localtime(fin).strftime("%d/%m/%Y %H:%M:%S")

    encabezados = ["Fecha", "Cámara", "Persona detectada", "Estado EPP", "Descripción"]

    ws.append([])
    ws.append(encabezados)

    fila_encabezado = 4

    for col in range(1, len(encabezados) + 1):
        celda = ws.cell(row=fila_encabezado, column=col)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(
            start_color="386FA4",
            end_color="386FA4",
            fill_type="solid"
        )
        celda.alignment = Alignment(horizontal="center")

    for informe in informes:
        ws.append([
            timezone.localtime(informe.fecha).strftime("%d/%m/%Y %H:%M:%S"),
            informe.camara,
            informe.persona_detectada or "No identificada",
            estado_epp(informe),
            informe.descripcion or ""
        ])

    # Ajuste de ancho de columnas corregido
    anchos = {
        "A": 22,
        "B": 25,
        "C": 30,
        "D": 20,
        "E": 60,
    }

    for columna, ancho in anchos.items():
        ws.column_dimensions[columna].width = ancho

    # Permite que la descripción no se corte visualmente
    for fila in ws.iter_rows(min_row=5, max_row=ws.max_row):
        for celda in fila:
            celda.alignment = Alignment(vertical="top", wrap_text=True)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}.xlsx"'

    return response


def generar_reporte_pdf(informes, nombre_archivo, nombre_periodo, inicio, fin):
    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )

    styles = getSampleStyleSheet()
    elementos = []

    titulo = Paragraph(f"<b>Reporte {nombre_periodo} de Informes</b>", styles["Title"])
    rango = Paragraph(
        f"Desde: {timezone.localtime(inicio).strftime('%d/%m/%Y %H:%M:%S')} "
        f" - Hasta: {timezone.localtime(fin).strftime('%d/%m/%Y %H:%M:%S')}",
        styles["Normal"]
    )

    elementos.append(titulo)
    elementos.append(rango)
    elementos.append(Spacer(1, 12))

    data = [["Fecha", "Cámara", "Persona", "Estado EPP", "Descripción"]]

    for informe in informes:
        data.append([
            timezone.localtime(informe.fecha).strftime("%d/%m/%Y %H:%M:%S"),
            informe.camara,
            informe.persona_detectada or "No identificada",
            estado_epp(informe),
            Paragraph(informe.descripcion or "", styles["BodyText"])
        ])

    if not informes.exists():
        data.append(["Sin registros", "-", "-", "-", "No existen informes en este período."])

    tabla = Table(
        data,
        repeatRows=1,
        colWidths=[100, 100, 120, 100, 320]
    )

    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#386FA4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F6F9FC")]),
    ]))

    elementos.append(tabla)
    doc.build(elementos)

    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}.pdf"'

    return response
